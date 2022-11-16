# -*- coding: utf-8 -*-
"""
Created on Fri Oct 14 11:38:45 2022

@author: user
"""

from loadexp import Dataloads, fileloads, build_data, GUI_load
import os
# from loadexp_0318 import *
from openpyxl import load_workbook
import csv
from pathlib import Path, PureWindowsPath
import sys
import pandas as pd
# import peakdetect




# year_path  = "D:\\Researcher\\JYCheon\\DATA\\Electrochemistry"

class Wonatech_tot(Dataloads):
    def __init__(self, path, file):
        Dataloads.__init__(self, path, file)
        
        print(f'Loading and splitting {self.file} file....................')
        self.wb = load_workbook(self.file_path)
        self.sheet_count = len(self.wb.sheetnames)
        self.raw_split = os.path.join(path, 'raw_split\\')
        
        if not os.path.exists(self.raw_split):
            os.mkdir(self.raw_split)
        
        for i in range(0, self.sheet_count, 2):
            name = self.wb[self.wb.sheetnames[i]]["A2"].value.split('\\')[-1].replace('.wrd', "")                                                                      
            ws = self.wb[self.wb.sheetnames[i+1]]
            
            with open(f'{self.raw_split}{name}.csv', 'w', newline = "", encoding = "cp949") as f:
                csv_writer = csv.writer(f)
                for r in ws.iter_rows():
                # for r in tqdm(ws.iter_rows()):
                    csv_writer.writerow([cell.value for cell in r])
                    
                    
# def main(date_path = year_path):
    
#     raw, path, _, _ = fileloads(date_path, "xlsx")
    
#     exp_obj = build_data(path, raw, Wonatech_tot)
    

# if __name__ == "__main__":
#     main()



path = Path(os.getcwd())
path2 = PureWindowsPath(os.getcwd())

path_ref = "path_ref.pkl"
xl_ref = "path_ref.xlsx"

pkl_path = path.parent.parent.joinpath(path_ref)
xl_path = path.parent.parent.joinpath(xl_ref)

df = pd.read_pickle(pkl_path)