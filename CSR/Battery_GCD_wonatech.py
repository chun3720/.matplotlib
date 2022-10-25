# -*- coding: utf-8 -*-
"""
Created on Wed Mar 24 09:06:16 2021
last update : Feb 15 2022
@author: jycheon
"""
from loadexp import Dataloads, fileloads, build_data
import pandas as pd
import os
import numpy as np
from matplotlib import pyplot as plt
import openpyxl
from pathlib import Path
from tqdm import tqdm

# plt.style.use('science')
plt.style.use(['science', 'no-latex'])
year_path = "D:\\Researcher\\JYCheon\\DATA\\Electrochemistry\\Coin cell\\2022"

class LIB_tot(Dataloads):   
    
    def __init__(self, path, file):       
        Dataloads.__init__(self, path, file)
        self.raw = pd.read_pickle(self.file_path)
        self.X, self.Y = self.raw.columns[0], self.raw.columns[1]
        self.null = self.raw[self.raw[self.X] == 0].index
        self.raw = self.raw.drop(self.null)
        self.raw =  self.raw.reset_index(drop = True).drop(index = 0)
        fcols = self.raw.select_dtypes('float').columns
        self.raw[fcols] = self.raw[fcols].apply(pd.to_numeric, downcast='float')
        self.raw.index = self.raw.index.astype('int32')
        
        self.electrode = 'c'
        self.cut = self.raw.iloc[50:, 1].min()
    
        
    def get_check(self):
        plt.plot(self.raw.index, self.raw[self.Y], label = "potential profile")
        plt.xlabel("Index")
        plt.ylabel("Potential (V vs. Li/Li$^+$)")
        plt.show()
        
        
        if self.raw.iloc[0, 1] > self.raw.iloc[5, 1]:
            self.electrode = "a"
            self.cut = self.raw.iloc[50:, 1].max()
        
        return self.electrode, self.cut
        
        
    def separation(self, electrode, cut_off):
        if electrode == 'a':
            if self.raw.iloc[0, 1] > cut_off:
                null = self.raw[self.raw[self.Y] < cut_off].index 
                self.data = (
                    self.raw.iloc[null[0]:]
                    .reset_index(drop = True)
                    .drop(index = 0)
                    )

            else:
                self.data = self.raw
            self.idx_list_values = self.raw[self.raw[self.Y] > cut_off].index
            
        
        elif electrode =='c':
            if self.raw.iloc[0, 1] < cut_off:
                null = self.raw[self.raw[self.Y] > cut_off].index
                self.data = (
                    self.raw.iloc[null[0]:]
                    .reset_index(drop = True)
                    .drop(index = 0)
                    )
                  
            else:
                self.data = self.raw
            self.idx_list_values = self.data[self.data[self.Y] < cut_off].index
        
        self.idxx = [0] + list(self.idx_list_values)
        output_path = os.path.join(self.path, 'split\\')
        
        if not os.path.exists(output_path):
            os.mkdir(output_path)
        n = len(self.idxx)-1
        idx_list, capacity_list = [], []
        # progress_bar(0, n)
        
        for i in tqdm(range(n)):
            df = self.data.iloc[self.idxx[i]:self.idxx[i+1]]
            # df.to_csv(f'{output_path}cycle{i+1}.csv', encoding = "cp949")
            df.to_parquet(f'{output_path}cycle{i+1}.pqt')
            k = df.shape[0]
            cap, vol = df.iloc[k-1]
            capacity_list.append(cap)
            idx_list.append(i+1)
            # progress_bar(i+1, n)     
                 
        d = {"Cycle": idx_list, "Capacity (Ah/g)" : capacity_list}
        df1 = pd.DataFrame(data = d, index = idx_list)      
        
        df1.plot(x = "Cycle", y = "Capacity (Ah/g)", kind = "scatter")
        plt.show()
        
        cycle_path = (
            Path(self.path)
            .parent
            .joinpath("cycle_auto")
            )

        if not cycle_path.exists():
            cycle_path.mkdir()
        
        target_file = cycle_path.joinpath(f'{self.name}_cycle.xlsx')
        with pd.ExcelWriter(target_file) as writer:
            df1.to_excel(writer, sheet_name = '데이터_1_1', index = False)     
    
        return output_path
    

    def __str__(self):
        self.label = self.name.replace("_", "-")
        return self.label
        
class LIB_csv(Dataloads):  
    def __init__(self, path, file):       
        Dataloads.__init__(self, path, file)
        self.df = pd.read_parquet(self.file_path) 
        self.X, self.Y  = self.df.columns[0], self.df.columns[1]
        self.null = self.df[self.df[self.X] == 0].index
        self.df = self.df.drop(self.null)
        self.min = self.df[self.Y].idxmin()
        self.max = self.df[self.Y].idxmax()
        self.positive = None
        self.negative = None
    
    def __str__(self):
        self.label = self.name.replace("_", "-")
        return self.label
    
    def df_indexing(self):
        
        if self.df.iloc[0, 1] > self.df.iloc[1, 1]:
            #case for anode
            self.negative = self.df.loc[:self.min] # charge
            self.positive = self.df.loc[self.min+1 :]   #discharge
            
            # bot = self.positive[self.X].idxmin()
            
            # c, v = self.positive.loc[bot]
            
            # if self.positive.iloc[0, 0] > c:
            #     self.positive = self.positive.loc[bot:]
        
        else:
            #case for cathode
            self.positive = self.df.loc[:self.max] #charge
            self.negative = self.df.loc[self.max +1 :] #discharge
            
            # top = self.negative[self.X].idxmin()
            # c, v = self.negative.loc[top]
            
            # if self.negative.iloc[0, 0] > c:
            #     self.negative = self.negative.loc[top:]
            
            
    def get_name(self):
        return self.name
              
    def get_discharge(self):
        return self.df.loc[self.max]
        
    def get_charge(self):
        return self.df.loc[self.min]
    
    def GCD_plot(self, color):
        plt.plot(self.negative[self.X], self.negative[self.Y], color, label = LIB_csv.__str__(self))
        plt.plot(self.positive[self.X], self.positive[self.Y], color )


def get_plot(path, exp_obj):
    color_list = ['k', 'r', 'tab:orange', 'g', 'b', 'purple', 'dimgrey', 'hotpink', 'steelblue', 'mediumseagreen', 'm']
    # n = len(exp)
    # nc = len(color_list)
    # numbering = range(k-1, n, k)
    n = len(color_list)
    for i, exp in enumerate(exp_obj):
        exp.df_indexing()
        exp.GCD_plot(color_list[i%n])
        # print(exp.min, exp.max)
    
    plt.xlabel('Specific Capacity (Ah/g)')
    plt.ylabel(r'Potential (V vs. Li/Li$^+$)')  
    plt.show()
        
        
    # for ix, num in enumerate(numbering):
    #     exp[num].df_indexing()
    #     exp[num].GCD_plot(color_list[ix%nc])
    
    # leg = plt.legend(fontsize = 6)
    # for line, text in zip(leg.get_lines(), leg.get_texts()):
    #     text.set_color(line.get_color())
    # plt.xlabel('Specific Capacity (Ah/g)')
    # plt.ylabel(r'Potential (V vs. Li/Li$^+$)')  
    # plt.savefig(path + '_GCD.png', dpi=300)

# def get_result(path, exp):
#     n = len(exp)
#     numbering = range(k-1, n, k)
#     for i in numbering:
#         print(exp[i].get_name().rjust(42))
#         print(exp[i].get_charge())
#         print('------------------------------------------')
#         print(exp[i].get_discharge(), end='\n')
#         print('==========================================')       
#     print('\nDone!\n')   


def get_capacity_df(exp_obj):
    return pd.concat([exp_obj.negative.reset_index(drop = True)
               , exp_obj.positive.reset_index(drop = True)], axis = 1)
    

def get_export(path, exp_obj):
    output_path = os.path.join(path, 'output\\')
    # output_path = self.path + 'output\\'
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    # GCDs = []
    # n = len(exp_obj)
    # numbering = range(k-1, n, k)
    # for num, b in zip(numbering, exp_obj):
        # with pd.ExcelWriter(f'{output_path}{exp_obj[num].name}_corrected.xlsx') as writer:
        # # with pd.ExcelWriter(output_path + exp[a].get_name() + '_corrected.xlsx') as writer:
        #     (
        #         pd.concat([exp_obj[num].negative.reset_index(drop = True)
        #                    ,exp_obj[num].positive.reset_index(drop = True)]
        #                    ,axis = 1, ignore_index = True)
        #                    .to_excel(writer, index = False
        #                              ,header = ["negative (Ah/g)","V1", "positive (Ah/g)", "V2"])                   
        #         )       
        # GCDs.append(exp_obj[num])
    
    pkl_file = f'{output_path}total.pkl'
    # with pd.ExcelWriter(f'{output_path}total.xlsx') as writer:
    # num = len(exp_obj)
    # progress_bar(0, num)
    tot_df = pd.DataFrame()
    header_tot = []
    for ix, gcd in enumerate(exp_obj):
        header = [f"negative_{ix}", f"V_{ix}", f"positive_{ix}", gcd.name]
        cap_df = get_capacity_df(gcd)
        # (
        #     cap_df
        #     .to_excel(writer, startcol = 4*ix, index = False
        #               , header = header, engine = "xlsxwriter")
        #     )
        tot_df = pd.concat([tot_df, cap_df], axis = 1, ignore_index  = True)
        header_tot += header
        # progress_bar(ix+1, num)
        
    tot_df.columns = header_tot
    tot_df.to_pickle(pkl_file)
            
            
def csv_from_excel(path, file):
    
    print("loading............................\n")
    wb = openpyxl.load_workbook(path + file, read_only = True, data_only = True)
    sheet_names = wb.worksheets
    target_sheet = sheet_names[1].title
    # ws = wb.get_sheet_by_name(target_sheet)
    ws = wb[target_sheet]
    name, ext = os.path.splitext(file)
    
    # with open(path + name + '.csv', 'w', newline = "") as f:
    #     csv_writer = csv.writer(f)
    #     for r in tqdm(ws.iter_rows()):            
    #         csv_writer.writerow([cell.value for cell in r])
    temp = []
    for r in tqdm(ws.iter_rows()):
        temp.append([cell.value for cell in r])
    
    df = pd.DataFrame(np.array(temp[1:]), columns = temp[0])
    fcols = df.select_dtypes('float').columns
    icols = df.select_dtypes('integer').columns

    df[fcols] = df[fcols].apply(pd.to_numeric, downcast='float')
    df[icols] = df[icols].apply(pd.to_numeric, downcast='integer')
    df.set_index("인덱스").to_pickle(f'{path}{name}.pkl')

 
def main(year_path):
    done0 =  False    
    
    while not done0:
        get_input = input("select data type, xlsx (x) or parquet (p): ")    
        if get_input in ["X", "x", "P", "p"]:
            done0 = True
        else:
            print("typing error!")
    
    if get_input.lower() == "x":
        ext = 'xlsx'
    elif get_input.lower() == "p":
        ext = "pqt"
    
    raw, path, _, _ = fileloads(year_path, ext)
    
    # if get_input == "x" and len(raw) > 1:
    #     import Battery_GCDplot_old as xl
    #     exp_obj = build_data(path, raw, xl.LIB_builder)
    #     xl.get_plot(path, exp_obj)
    #     xl.get_result(path, exp_obj)
    #     xl.get_export(path, exp_obj)
        
    if get_input == "p" and len(raw) > 1:
        exp_obj= build_data(path, raw, LIB_csv)
        get_plot(path, exp_obj)
        # get_result(path, exp_obj, k=1)
        get_export(path, exp_obj)
      
    
    else:
    
        if (os.path.splitext(raw[0])[0] + ".pkl") not in os.listdir(path):
            csv_from_excel(path, raw[0])
            csv_list = [_ for _ in os.listdir(path) if _.endswith(".pkl")]
        else:
            csv_list = [_ for _ in os.listdir(path) if _.endswith(".pkl")]
            
        
        exp_data = build_data(path, csv_list, LIB_tot)
        electrode, cutoff = exp_data[0].get_check()
        
        print(electrode, cutoff)
        
        # print(exp_data[0].electrode, exp_data[0].cut)
        # # done1 = False
        # while not done1:
        #     electrode = input("which electrode? a, anode or c, cathode (or full cell): ")
            
        #     if electrode in ['A', 'a', 'c', "C"]:
        #         done1 = True
        #     else:
        #         print("typing error!")         
            
        # electrode = electrode.lower()
        # done2 =  False
        # while not done2:
        #     cutoff = input("type cutoff voltage for each cycle (anode-> top, cathode (or full cell) -> bottom): ")
        #     if cutoff[0].isnumeric():
        #         done2 = True
        #     else:
        #         print("typing error!")
                
        
        output_path = exp_data[0].separation(electrode, round(cutoff, 1))
        
        raw_list = os.listdir(output_path)
        raw_list = [_ for _ in raw_list if _.endswith(".pqt")]
        sorted_list = sorted(raw_list, key = len)
        exp_obj = build_data(output_path, sorted_list , LIB_csv)
        
        # done3 = False
        # while not done3:
            
        #     k = input("\ntype separation unit: ")
            
        #     if k[0].isnumeric():
        #         done3 = True
        #     else:
        #         print("typing error!")
        # k = int(k)
        def final_plot(output_path, exp_obj):
            get_plot(output_path, exp_obj)
            # get_result(output_path, exp_obj)
            get_export(output_path, exp_obj)
                   

        final_plot(output_path, exp_obj)
   

if __name__ == "__main__":
    # year_path = get_data_folder(Path(sys.argv[0]).name)
    main(year_path)
    
    
    