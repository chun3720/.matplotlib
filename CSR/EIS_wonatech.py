# -*- coding: utf-8 -*-
"""
Created on Wed Mar 30 14:26:05 2022

@author: user
"""

from loadexp import Dataloads, fileloads, build_data, get_data_folder
import pandas as pd
import os
from matplotlib import pyplot as plt
# from loadexp_0318 import *
from openpyxl import load_workbook
import csv

# import peakdetect


plt.style.use(['science', 'no-latex'])
year_path  = "D:\\Researcher\\JYCheon\\DATA\\Electrochemistry\\EIS"


axis = 1000


class EIS_tot(Dataloads):
    def __init__(self, path, file):
        Dataloads.__init__(self, path, file)
        
        print(f'Loading and splitting {self.file} file....................')
        self.wb = load_workbook(self.file_path)
        self.sheet_count = len(self.wb.sheetnames)
        self.raw_split = os.path.join(path, 'raw_split')
        
        if not os.path.exists(self.raw_split):
            os.mkdir(self.raw_split)
        
        for i in range(0, self.sheet_count, 2):
            name = self.wb[self.wb.sheetnames[i]]["A2"].value.split('\\')[-1].replace('.SEO', "")                                                                      
            ws = self.wb[self.wb.sheetnames[i+1]]
            
            file2export = os.path.join(self.raw_split, f"{name}.csv")
            with open(file2export, 'w', newline = "", encoding = "cp949") as f:
                csv_writer = csv.writer(f)
                for r in ws.iter_rows():
                # for r in tqdm(ws.iter_rows()):
                    csv_writer.writerow([cell.value for cell in r])
            
            
        # print("done!")
        
class EIS_raw(Dataloads):
    def __init__(self, path, file):
        Dataloads.__init__(self, path, file)
        
        self.raw = pd.read_csv(self.file_path, index_col = 0, encoding = "cp949")
        self.output_path = os.path.join(path, "output")
        # self.output_path = path + 'output\\'
        if not os.path.exists(self.output_path):
            os.mkdir(self.output_path)
        
        
def raw_plot(path, obj_list):
    color_list = ['k', 'r', 'tab:orange', 'g', 'b', 'm', 'gray', 'brown','darkcyan', 
                  'skyblue', 'hotpink', 'dodgerblue']
    n = len(color_list)
    
    x, y = obj_list[0].raw.columns
    file2export = os.path.join(obj_list[-1].output_path, "EIS_total.xlsx")
    with pd.ExcelWriter(file2export) as writer:
        for i, exp in enumerate(obj_list):
            plt.plot(exp.raw[x], exp.raw[y], label = exp.name, color = color_list[i%n])
            (exp.raw[[x, y]]
             .to_excel(writer, startcol = 2*i, index = False, header = [f'{i}', exp.name] )
             
             )

    leg = plt.legend(fontsize = "xx-small")
    for line, text in zip(leg.get_lines(), leg.get_texts()):
        text.set_color(line.get_color())
    plt.xlim(0, axis)
    plt.ylim(0, axis)
    plt.xlabel("$Z'[Ohms]$", fontsize = 12)
    plt.ylabel("$-Z''[Ohms]$", fontsize = 12)
    plt.show()



def main(py_path):    
    done = False

    while not done:
        check = input("select method: splitting total data(x) or not (c): ")
        
        
        if check.lower() == "x":
            done = True
            
            raw, path, _, _ = fileloads(py_path, ".xlsx")
            path2check = os.path.join(path, "raw_split")
            if not os.path.exists(path2check):
                
                exp_obj = build_data(path, raw, EIS_tot)
            
            raw_path = path2check
            raw_list = [_ for _ in os.listdir(raw_path) if _.endswith(".csv")]

            sep_obj = build_data(raw_path, raw_list, EIS_raw)
            raw_plot(raw_path, sep_obj)
            
        elif check.lower() == "c":
            
            done = True
            
            raw, path, _, _ = fileloads(py_path, ".csv")
            sep_obj = build_data(path, raw, EIS_raw)
            raw_plot(path, sep_obj)
            
        else:
            print("Unvalid option!!\n")


if __name__ == "__main__":
    main(year_path)