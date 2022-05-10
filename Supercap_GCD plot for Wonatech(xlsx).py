# -*- coding: utf-8 -*-
"""
Created on Wed Mar 30 14:26:05 2022

@author: user
"""

from loadexp import *
import pandas as pd
import os
from matplotlib import pyplot as plt
import numpy as np
# from loadexp_0318 import *
from openpyxl import load_workbook
import csv
import tqdm
# import peakdetect


plt.style.use(['science', 'no-latex'])
year_path  = "D:\\Researcher\\JYCheon\\DATA\\Electrochemistry\\2022\\Raw"

class Supercap(Dataloads):
    def __init__(self, path, file):
        Dataloads.__init__(self, path, file)
        
        print(f'Loading {self.file} file....................')
        self.wb = load_workbook(self.file_path)
        self.sheet_count = len(self.wb.sheetnames)
        self.raw_split = os.path.join(path, 'raw_split\\')
        
        if not os.path.exists(self.raw_split):
            os.mkdir(self.raw_split)
        
        for i in range(0, self.sheet_count, 2):
            name = self.wb[self.wb.sheetnames[i]]["A2"].value.split('\\')[-1].replace('.wrd', "")                                                                      
            ws = self.wb[self.wb.sheetnames[i+1]]
            
            with open(f'{self.raw_split}{name}.csv', 'w', newline = "") as f:
                csv_writer = csv.writer(f)
                for r in ws.iter_rows():
                # for r in tqdm(ws.iter_rows()):
                    csv_writer.writerow([cell.value for cell in r])
            
            
        # print("done!")

class Capacitance(Dataloads):
    def __init__(self, path, file):
        Dataloads.__init__(self, path, file)
        self.raw = pd.read_csv(self.file_path, header = 0, index_col = 0, sep = ",", encoding = "cp949")
        self.raw.columns = ["T", "V", "I"]
        self.raw[["day", "hour", "min", "sec" ]] = self.raw["T"].str.split(":", expand = True).astype(float)
        self.raw["time"] = self.raw["min"]*60 + self.raw["sec"]
        self.cols = ["time", "V", "I"]
        self.df = self.raw[self.cols]

        # self.df = pd.read_csv(self.file_path, index_col = 0, encoding = "cp949", header = 0)
        
        self.null = self.df[self.df["I"] == 0].index
        self.df = self.df.drop(self.null)
        # self.df = self.df.iloc[5:]
        self.df.reset_index(drop = True, inplace = True)
        self.idx_list = self.df[self.df["V"] < 0].index
        self.max_point = None
        self.half_point = None
        # print(list(self.idx_list))
        # print(len(self.idx_list))
        if len(self.idx_list) != 4:
            print(f'{self.file} may have some error please check')
        self.appl_current = abs(self.df.iloc[0, 2]) # A
        self.appl_unit = 'A'
        
        if self.appl_current < 1 and self.appl_unit == 'A':
            self.appl_current *= 1000
            self.appl_unit = 'mA'
            
        if self.appl_current < 0.1 and self.appl_unit == 'mA':
            self.appl_current *= 1000
            self.appl_unit = 'uA'
        
        self.Is = abs(self.df.iloc[0, 2]) # A
        self.input = abs(self.Is) * 1000000 # uA
        self.idxx = list(self.idx_list)
        self.cycle2 = self.df[self.cols].iloc[self.idxx[1]:self.idxx[2]].reset_index(drop = True)
        # self.cycle2.reset_index(drop = True, inplace = True)
        self.origin = self.cycle2["time"].loc[0]
        self.cycle2["time"] -= self.origin
        if self.cycle2["V"].loc[0] < 0:
            self.cycle2.drop(index = 0).reset_index(drop = True, inplace = True)
            # self.cycle2.reset_index(drop = True, inplace = True)
       
        self.cap_result = 0
        self.cap_unit = 'F'
        

    def cap_plot(self, path):
        output_path = f'{path}\\output\\'
        
        if not os.path.exists(output_path):
            os.mkdir(output_path)
        
        plt.plot(self.cycle2["time"], self.cycle2["V"], '--', color = 'gray', label = self.name)
        cap_label = self.get_capacitance()

        try:
            plt.plot(self.max_point, self.half_point, 'r-', label = cap_label)
            
        except:
            pass

        leg = plt.legend(fontsize = 'small')
        for line, text in zip(leg.get_lines(), leg.get_texts()):
            text.set_color(line.get_color())
        plt.xlabel('Time (s)', fontsize = 13)
        plt.xticks(fontsize = 11)
        plt.ylabel('Voltage (V)', fontsize = 13)
        plt.savefig(f'{output_path}{self.name}.png')
        plt.yticks(fontsize = 11)
        
    def get_calculation(self):
        
        try:
            k = self.cycle2["V"].idxmax()
            T1, V1 = self.cycle2[["time", "V"]].loc[k]
            self.discharge = self.cycle2.loc[k:]
            if V1 > 2.4:
                idx_list = self.discharge[self.discharge["V"] < 1.5].index
                T2, V2 = self.cycle2[["time", "V"]].loc[idx_list[0]]
            else:
            ###
                # for 1V calculation
                n = self.cycle2.shape[0]
                # n = self.discharge.shape[0]
                T2, V2= self.cycle2[["time", "V"]].loc[n-1]
                # T2, V2 = self.discharge.loc[n-1]
    
            self.max_point = np.array([T1, T2])
            self.half_point = np.array([V1, V2])
            self.slope = (T2- T1) /  (V1-V2)
            self.cap_result = self.Is * self.slope
            return True
        except:
            return None
        
    def get_capacitance(self):

        if self.cap_result < 1 and self.cap_unit =='F':
            self.cap_result *= 1000
            self.cap_unit = 'mF'
            
        if self.cap_result < 1 and self.cap_unit == 'mF':
            self.cap_result *= 1000
            self.cap_unit = 'uF'
            
        res = f'{round(self.cap_result, 2)} {self.cap_unit}'
        return res
    
    def get_condition(self):
        
        return f'{round(self.appl_current, 2)} {self.appl_unit}'
        

def get_cap_result(exp_obj, path):
    for exp in exp_obj:
        exp.get_calculation()
        exp.cap_plot(path)
        plt.show()

def get_multiplot(exp_obj, path):
    
    for exp in exp_obj:
        plt.plot(exp.cycle2["time"], exp.cycle2["V"])
        
    plt.show()
    
def get_export(exp_obj, path):
    output_path = f'{path}\\output\\'
    
    if not os.path.exists(output_path):
        os.mkdir(output_path)
  
    GCD_list = []
    cap_list = []
    cap_unit = []
    Is_list = []
    
    for gcd in exp_obj:
        GCD_list.append(gcd.name)
        Is_list.append(gcd.get_condition())
        cap_list.append(round(gcd.cap_result, 2))
        cap_unit.append(gcd.cap_unit)
    
    d = {"Capacitance (I*dt/dV)": cap_list, "unit": cap_unit, "Current": Is_list}
    df1 = pd.DataFrame(data = d, index = GCD_list)
    
    # with pd.ExcelWriter(output_path + "GCD_tot.xlsx") as writer:
    with pd.ExcelWriter(f'{output_path}\\GCD_tot.xlsx') as writer:
        n = len(exp_obj)
        print("exporting............\n")
        progress_bar(0, n)
        for i, gcd in enumerate(exp_obj):
            cols = ["time", "V"]
            (
                gcd.cycle2[cols]
                .to_excel(writer, startcol = 2*i, index = False, header = [f'{i}', gcd.name])
                
                )

            progress_bar(i+1, n)
        df1.to_excel(writer, sheet_name = 'Summary')
        
    with pd.ExcelWriter(f'{output_path}Raw_tot.xlsx') as writer:
        n = len(exp_obj)
        progress_bar(0, n)
        for ix, exp in enumerate(exp_obj):
            cols = ["time", "V"]
            (
                exp.raw[cols].to_excel(writer, startcol = 2*ix, index = False, header = [f'{ix}, {exp.get_condition()}', exp.name])
                )
            progress_bar(i+1, n)
        
        
raw, path, _, _ = fileloads(year_path, ".xlsx")
if not os.path.exists(f'{path}raw_split\\'):
    
    exp_obj = build_data(path, raw, Supercap)
    
#     exp_obj[0].get_sep()


# exp_path = path + 'raw_split\\'
exp_path = f'{path}raw_split\\'
exp_list = [_ for _ in os.listdir(exp_path) if _.endswith(".csv") ]

sep_obj = build_data(exp_path, exp_list, Capacitance)


get_cap_result(sep_obj, exp_path)
get_multiplot(sep_obj, exp_path)
get_export(sep_obj, exp_path)
